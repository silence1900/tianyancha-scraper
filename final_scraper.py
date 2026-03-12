"""
天眼查企业信息爬虫 v4
思路：Google 搜索 "公司名 site:tianyancha.com" → 获取天眼查详情页 URL → 连接你手动打开的 Chrome 访问

启动步骤：
1. 关闭所有 Chrome 窗口
2. 用命令行启动 Chrome（开启远程调试）：
   macOS:
     /Applications/Google\ Chrome.app/Contents/MacOS/Google\ Chrome --remote-debugging-port=9222 --no-first-run
   Windows:
     "C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --no-first-run
3. 在 Chrome 里手动登录天眼查（可选，不登录也能看到基本信息）
4. 运行此脚本：python scraper_v4.py

依赖：pip install playwright openpyxl
"""

import asyncio
import json
import re
import random
from pathlib import Path
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 配置
# ============================================================
INPUT_FILE = "files/企业信息查询表.xlsx"
OUTPUT_FILE = "files/企业信息查询结果_final.xlsx"
PROGRESS_FILE = "files/progress_final.json"
CDP_PORT = 9222
TEST_LIMIT = None

NON_LEGAL_ENTITIES = ["杭州弧途科技-用工业务部"]

# ============================================================
# 人工介入逻辑
# ============================================================
def wait_for_human_intervention(reason: str):
    """暂停脚本，等待人工介入"""
    print("\n" + "!" * 60)
    print(f"  ⚠️  检测到需要人工介入：{reason}")
    print("  👉 请在 Chrome 浏览器中手动完成验证（如滑块、验证码）或刷新页面")
    print("  ✅ 完成后，请回到此处按回车键继续...")
    print("!" * 60 + "\n")
    # 播放提示音（仅 macOS 有效，其他系统忽略）
    print("\a") 
    input()
    print("  ▶️  继续运行...")

# ============================================================
# 读取 Excel 中的企业列表
# ============================================================
def read_companies_from_excel(file_path):
    companies = []
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 从第3行开始读取，第2列是企业名称
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[1]: # 确保企业名称不为空
                company_name = str(row[1]).strip()
                if company_name not in companies: # 简单去重
                    companies.append(company_name)
    except Exception as e:
        print(f"读取 Excel 文件失败: {e}")
        return []
    return companies


# ============================================================
# 第一步：用天眼查站内搜索找到详情页 URL
# ============================================================

async def find_tianyancha_url(page, name: str) -> str:
    """通过天眼查站内搜索获取详情页链接"""
    search_url = f"https://www.tianyancha.com/search?key={name}"
    
    try:
        await page.goto(search_url, wait_until="domcontentloaded", timeout=20000)
        await asyncio.sleep(random.uniform(2, 3))

        # 检查是否触发验证码/拦截
        if "验证" in await page.title() or await page.query_selector(".baxia-dialog") or "反爬" in await page.inner_text("body"):
            wait_for_human_intervention("触发天眼查验证/反爬拦截")
            # 人工处理后，重新加载页面
            await page.reload(wait_until="domcontentloaded")
            await asyncio.sleep(2)

        # 获取搜索结果列表中的第一个公司
        # 通常结构是 .search-item 或 .result-list 里的第一个 a 标签
        # 这里的选择器需要根据实际页面结构调整，先尝试通用的
        
        # 尝试获取第一个搜索结果的名称和链接
        # 天眼查搜索结果标题通常在 .header a 或 .name a 中
        first_result = await page.query_selector(".index_list-wrap___axcs .index_name__qEdWi a")
        
        if not first_result:
             # 备用选择器
            first_result = await page.query_selector("a[href*='company/']")
            
        if first_result:
            href = await first_result.get_attribute("href")
            text = await first_result.inner_text()
            
            # 简单的名称清洗（去空格、括号统一）
            clean_target = name.replace("（", "(").replace("）", ")").strip()
            clean_found = text.replace("（", "(").replace("）", ")").strip()
            
            # 只有当名称包含目标名称，或者高度相似时才采纳
            # 严格模式：要求前几个字必须匹配
            if clean_target[:4] in clean_found:
                 if href and not href.startswith("http"):
                    href = "https://www.tianyancha.com" + href
                 return href
            else:
                print(f"    ⚠️ 搜索结果不匹配: 目标[{name}] vs 结果[{text}]")
                return ""

    except Exception as e:
        print(f"    站内搜索失败: {e}")

    return ""


# ============================================================
# 第二步：访问天眼查详情页提取数据
# ============================================================

async def extract_from_tianyancha(page, url: str, name: str) -> dict:
    result = {"name": name, "credit_code": "", "registered_capital": "",
              "registration_date": "", "status": "", "remark": ""}
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=25000)
        await asyncio.sleep(random.uniform(2, 3))

        # 检查是否需要登录弹窗（天眼查基本信息不登录也能看）
        login_modal = await page.query_selector(".dialog-content, .login-dialog, [class*='login-modal']")
        if login_modal:
            # 尝试关闭弹窗
            close_btn = await page.query_selector(".close-btn, .dialog-close, [class*='close']")
            if close_btn:
                await close_btn.click()
                await asyncio.sleep(1)

        text = await page.inner_text("body")

        # 统一社会信用代码（18位字母数字）
        m = re.search(r"统一社会信用代码[：:\s]*([A-Z0-9]{18})", text)
        if m:
            result["credit_code"] = m.group(1)

        # 注册资本 (尝试多种格式)
        # 格式示例: "注册资本： 1000万人民币" 或 "注册资本：1000万元"
        capital_patterns = [
            r"注册资本[：:\s]*([0-9,.]+\s*(?:万|亿|万美元|万港元|万人民币|万元人民币|万元|元)[人民币]*)",
            r"注册资本[：:\s]*([0-9,.]+\s*[^\s<]+)", # 宽泛匹配
        ]
        for pat in capital_patterns:
            m = re.search(pat, text)
            if m:
                # 清理提取到的内容
                cap = m.group(1).strip()
                # 过滤掉显然不对的长文本
                if len(cap) < 30:
                    result["registered_capital"] = cap
                    break

        # 成立日期 / 注册时间
        for pat in [r"成立日期[：:\s]*(\d{4}-\d{2}-\d{2})",
                    r"注册时间[：:\s]*(\d{4}-\d{2}-\d{2})",
                    r"成立时间[：:\s]*(\d{4}-\d{2}-\d{2})"]:
            m = re.search(pat, text)
            if m:
                result["registration_date"] = m.group(1)
                break

        # 经营状态
        m = re.search(r"(存续|注销|吊销|迁出|撤销|开业|在营|登记)", text)
        if m:
            result["status"] = m.group(1)
            # 如果是注销/吊销，添加到备注
            if result["status"] in ["注销", "吊销", "撤销"]:
                result["remark"] = f"⚠️ 企业已{result['status']}" + (" | " + result["remark"] if result["remark"] else "")

        # 验证：名称是否匹配（防止搜到错误公司）
        # 取页面 title 或 h1 做校验
        title = await page.title()
        clean_name = name.replace("（", "(").replace("）", ")")
        page_name  = title.replace("（", "(").replace("）", ")")
        if name[:4] not in title and clean_name[:4] not in page_name:
            result["remark"] = f"⚠️ 页面公司名可能不匹配，请核查（页面：{title[:20]}）"

    except Exception as e:
        result["remark"] = f"提取出错: {str(e)[:60]}"
        print(f"    ❌ 提取失败: {e}")

    return result


# ============================================================
# 主查询函数
# ============================================================

async def query_company(page, name: str) -> dict:
    result = {"name": name, "credit_code": "", "registered_capital": "",
              "registration_date": "", "status": "", "remark": ""}

    print(f"  🔍 搜索天眼查链接...")
    url = await find_tianyancha_url(page, name)

    if not url:
        result["remark"] = "未找到天眼查页面"
        print(f"  ❌ 未找到链接")
        return result

    print(f"  🌐 访问: {url}")
    result = await extract_from_tianyancha(page, url, name)
    result["name"] = name  # 保留原始名称

    if result["credit_code"]:
        print(f"  ✅ {result['credit_code']} | {result['registered_capital']} | {result['registration_date']}")
    else:
        if not result["remark"]:
            result["remark"] = "未提取到信用代码，请手动核查"
        print(f"  ⚠️  数据不完整")

    return result


# ============================================================
# Excel 输出
# ============================================================

def save_excel(results: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "企业信息"

    hf = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    af = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
    wf = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    gf = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    bd = Border(
        left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),  bottom=Side(style="thin", color="BFBFBF"),
    )

    headers = ["序号", "企业名称", "统一社会信用代码", "注册资本", "注册时间", "经营状态", "备注"]
    widths  = [6, 40, 22, 16, 14, 10, 30]

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = hf
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = bd
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22

    row = 2
    # 非法人实体
    for nm in NON_LEGAL_ENTITIES:
        for c, v in enumerate([row-1, nm, "—", "—", "—", "—", "非独立法人（部门）"], 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font      = Font(name="Arial", size=10, color="808080", italic=True)
            cell.fill      = gf
            cell.alignment = Alignment(horizontal="center" if c != 2 else "left", vertical="center")
            cell.border    = bd
        row += 1

    for i, r in enumerate(results):
        fill = wf if r.get("remark") else (af if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF"))
        for c, v in enumerate([row-1, r["name"], r["credit_code"], r["registered_capital"],
                                r["registration_date"], r["status"], r.get("remark", "")], 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center" if c in [1,3,4,5,6] else "left", vertical="center")
            cell.border    = bd
        ws.row_dimensions[row].height = 18
        row += 1

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"\n📊 Excel 已保存：{OUTPUT_FILE}，共 {row-2} 行数据")


# ============================================================
# 主流程
# ============================================================

async def main():
    print("=" * 60)
    print("  企业信息批量爬虫 v4（天眼查版）")
    print("=" * 60)
    print("""
【启动步骤】
1. 关闭所有 Chrome 窗口，然后运行：

   macOS:
     /Applications/Google\\ Chrome.app/Contents/MacOS/Google\\ Chrome \\
       --remote-debugging-port=9222 --no-first-run

   Windows:
     "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" ^
       --remote-debugging-port=9222 --no-first-run

2. Chrome 打开后，登录天眼查（不登录也能查基本信息）
3. 回到此终端按回车开始
    """)
    # input("✅ 确认 Chrome 已启动，按回车继续...")

    # 读取 Excel
    print(f"正在读取 {INPUT_FILE} ...")
    all_companies = read_companies_from_excel(INPUT_FILE)
    if not all_companies:
        print("❌ 未能读取到企业列表")
        return

    # 截取前 10 个 (如果有 TEST_LIMIT)
    companies_to_process = [c for c in all_companies if c not in NON_LEGAL_ENTITIES]
    if TEST_LIMIT:
        companies_to_process = companies_to_process[:TEST_LIMIT]
        print(f"共读取 {len(all_companies)} 家，本次测试 {len(companies_to_process)} 家")
    else:
        print(f"共读取 {len(all_companies)} 家，准备全量查询")

    # 断点续爬
    progress = {}
    if Path(PROGRESS_FILE).exists():
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            progress = json.load(f)
    done        = set(progress.get("done", {}).keys())
    todo        = [c for c in companies_to_process if c not in done]
    all_results = list(progress.get("done", {}).values())
    print(f"待查：{len(todo)} 家 / 已完成：{len(done)} 家\n")

    if not todo:
        save_excel(all_results)
        return

    async with async_playwright() as p:
        browser = await p.chromium.connect_over_cdp(f"http://localhost:{CDP_PORT}")
        print(f"✅ 已连接 Chrome（端口 {CDP_PORT}）\n")

        contexts = browser.contexts
        page = contexts[0].pages[0] if (contexts and contexts[0].pages) else await browser.new_page()

        consecutive_failures = 0  # 连续失败计数

        for i, name in enumerate(todo, 1):
            # 每 40 条强制暂停，提醒人工检查
            if i > 0 and i % 40 == 0:
                wait_for_human_intervention("已连续查询 40 条，天眼查可能会限制访问，建议手动刷新页面或验证")

            print(f"\n[{i}/{len(todo)}] {name}")
            result = await query_company(page, name)
            
            # 检查结果状态
            if not result["credit_code"] and "未找到" in result.get("remark", ""):
                consecutive_failures += 1
                if consecutive_failures >= 3:
                    wait_for_human_intervention("连续 3 次未找到有效信息，可能已被封控或需要验证")
                    consecutive_failures = 0  # 重置计数
            else:
                consecutive_failures = 0  # 成功一次就重置

            all_results.append(result)

            progress.setdefault("done", {})[name] = result
            with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
                json.dump(progress, f, ensure_ascii=False, indent=2)

            # 随机延迟，避免触发风控
            delay = random.uniform(3, 6)
            print(f"  ⏳ 等待 {delay:.1f}s")
            await asyncio.sleep(delay)

        await browser.close()

    name_map = {r["name"]: r for r in all_results}
    ordered  = [name_map[c] for c in companies_to_process if c in name_map]
    save_excel(ordered)
    print("\n🎉 全部完成！")


if __name__ == "__main__":
    asyncio.run(main())
